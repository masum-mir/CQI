import re
import urllib.parse

DAY_MAP = {'S': 'Sat', 'M': 'Mon', 'T': 'Tue', 'W': 'Wed', 'R': 'Thu', 'F': 'Fri', 'A': 'Fri'}

_TIMING = re.compile(
    r'^(?P<days>[SMTWRFA]{1,3})\s+'
    r'(?P<t1>\d{1,2}:\d{2}\s*[AP]M)\s*-\s*(?P<t2>\d{1,2}:\d{2}\s*[AP]M)\s*$',
    re.IGNORECASE)
_CODE = re.compile(r'^[A-Z]{2,4}\d{3,4}', re.IGNORECASE)

HEADER_KEYS = {
    'course': 'course', 'section': 'section', 'faculty': 'faculty',
    'timing': 'timing', 'room': 'room', 'department': 'department',
    'capacity': 'capacity', 'seat': 'seat',
}


def _to_24h(t):
    t = t.replace(' ', '')
    m = re.match(r'(\d{1,2}):(\d{2})(AM|PM)', t, re.IGNORECASE)
    if not m:
        return t
    hh, mm, ap = int(m.group(1)), m.group(2), m.group(3).upper()
    if ap == 'PM' and hh != 12:
        hh += 12
    if ap == 'AM' and hh == 12:
        hh = 0
    return f'{hh:02d}:{mm}'


def _expand_days(token):
    return [DAY_MAP.get(ch.upper(), ch) for ch in token]


def _looks_like_html(data):
    head = data[:2048].lstrip().lower()
    return head.startswith(b'<') or b'<table' in head or b'<html' in head


def _rows_from_html(data):
    try:
        from bs4 import BeautifulSoup
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError('beautifulsoup4 is required to parse Excel/HTML '
                           '(pip install beautifulsoup4 lxml)') from exc

    soup = BeautifulSoup(data, 'html.parser')
    table = soup.find('table') or soup
    trs = table.find_all('tr')

    colmap, start = None, 0
    for i, tr in enumerate(trs):
        texts = [c.get_text(strip=True).lower() for c in tr.find_all(['th', 'td'])]
        if any('course' in t for t in texts) and any('section' in t for t in texts):
            colmap = {}
            for idx, t in enumerate(texts):
                for key, kw in HEADER_KEYS.items():
                    if kw in t and key not in colmap:
                        colmap[key] = idx
            start = i + 1
            break

    out = []
    for tr in trs[start:]:
        tds = tr.find_all('td')
        if not tds:
            continue
        cells = [td.get_text(strip=True) for td in tds]
        title = semester = None
        a = tr.find('a', href=True)
        if a:
            q = urllib.parse.parse_qs(urllib.parse.urlparse(a['href']).query)
            if q.get('CourseName'):
                title = q['CourseName'][0].strip()
            if q.get('Semestername'):
                semester = q['Semestername'][0].strip()
        out.append({'cells': cells, 'colmap': colmap, 'title': title, 'semester': semester})
    return out


def _rows_from_xlsx(data):
    try:
        import io
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError('openpyxl is required for .xlsx files (pip install openpyxl)') from exc

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = [[('' if v is None else str(v)).strip() for v in r]
            for r in ws.iter_rows(values_only=True)]
    colmap, start = None, 0
    for i, cells in enumerate(rows):
        low = [c.lower() for c in cells]
        if any('course' in t for t in low) and any('section' in t for t in low):
            colmap = {}
            for idx, t in enumerate(low):
                for key, kw in HEADER_KEYS.items():
                    if kw in t and key not in colmap:
                        colmap[key] = idx
            start = i + 1
            break
    return [{'cells': c, 'colmap': colmap, 'title': None, 'semester': None} for c in rows[start:]]


def _col(cells, colmap, key, default_idx):
    idx = (colmap or {}).get(key, default_idx)
    return cells[idx].strip() if idx is not None and idx < len(cells) else ''


def parse_offered_courses(data, filename='upload.xls', departments=None):
    if hasattr(data, 'read'):
        data = data.read()

    ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
    if ext == 'xlsx' and not _looks_like_html(data):
        raw = _rows_from_xlsx(data)
    else:
        raw = _rows_from_html(data)

    dept_set = {d.upper() for d in departments} if departments else None
    offerings = {}
    errors = []
    semester = None

    for n, row in enumerate(raw, start=1):
        cells, colmap = row['cells'], row['colmap']
        code = _col(cells, colmap, 'course', 0).upper()
        if not _CODE.match(code):
            continue
        section = _col(cells, colmap, 'section', 1)
        faculty = _col(cells, colmap, 'faculty', 2).strip().upper()
        timing = _col(cells, colmap, 'timing', 3)
        room = _col(cells, colmap, 'room', 4)
        cap_total = _col(cells, colmap, 'capacity', 6)
        seat_taken = _col(cells, colmap, 'seat', 7)

        if row.get('semester') and not semester:
            semester = row['semester']

        dept = re.match(r'[A-Z]+', code).group()
        if dept_set and dept not in dept_set:
            continue

        key = (code, section)
        off = offerings.setdefault(key, {
            'course_code': code, 'section': section,
            'faculty_code': faculty or None,
            'title': row.get('title'),
            'capacity': None,
            'schedule': [],
        })
        if not off['faculty_code'] and faculty:
            off['faculty_code'] = faculty
        if not off['title'] and row.get('title'):
            off['title'] = row['title']

        def _int(x):
            m = re.search(r'\d+', x or '')
            return int(m.group()) if m else None
        total, enrolled = _int(cap_total), _int(seat_taken)
        if total is not None or enrolled is not None:
            off['capacity'] = {'enrolled': enrolled, 'total': total}

        m = _TIMING.match(timing)
        if m:
            for day in _expand_days(m.group('days')):
                off['schedule'].append({
                    'day': day,
                    'start_time': _to_24h(m.group('t1')),
                    'end_time': _to_24h(m.group('t2')),
                    'room': room,
                })
        elif timing and timing.upper() not in ('TBA', ''):
            errors.append({'row': n, 'message': f'Unrecognized timing "{timing}" for {code}-{section}'})

    return {'semester': semester, 'offerings': list(offerings.values()), 'errors': errors}